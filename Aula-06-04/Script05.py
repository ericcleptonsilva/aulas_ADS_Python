from openpyxl import Workbook
wb = Workbook()
# pegar a planilha ativa
ws = wb.active
# os dado podem ser atribuídos diretamente às células
ws['A1'] = "data de hoje"
# linhas também pode ser anexadas
ws.append([1,2,3])
# tipos de python serão convertido automaticamente
import datetime
ws['A2'] = datetime.datetime.now()
wb.save('exempl.xlsx')
print("feito")







